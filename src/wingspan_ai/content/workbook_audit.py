"""Audit helpers for the source Wingspan card workbook."""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from wingspan_ai.content.schemas import ContentPack

EXPECTED_SHEETS = ("Birds", "Hummingbirds", "Bonus", "Goals")

BIRD_COLUMNS = (
    "Common name",
    "Scientific name",
    "Set",
    "Color",
    "Power text",
    "Predator",
    "Flocking",
    "Bonus card",
    "Victory points",
    "Nest type",
    "Egg limit",
    "Wingspan",
    "Forest",
    "Grassland",
    "Wetland",
    "Invertebrate",
    "Seed",
    "Fish",
    "Fruit",
    "Rodent",
    "Nectar",
    "Wild (food)",
    "/ (food cost)",
    "* (food cost)",
    "Total food cost",
    "Beak direction",
    "Swift Start",
    "Automa ban",
)

BONUS_COLUMNS = (
    "Bonus card",
    "Set",
    "Automa",
    "Condition",
    "VP",
    "Explanatory text",
    "%",
)

GOAL_COLUMNS = ("Goal", "Set", 1, 2, 3, 4, "Reverse")

HUMMINGBIRD_COLUMNS = (
    "Common name",
    "Scientific name",
    "Group",
    "Benefit",
    "Beak direction",
)

REQUIRED_COLUMNS = {
    "Birds": (
        "Common name",
        "Scientific name",
        "Set",
        "Victory points",
        "Egg limit",
        "Wingspan",
        "Total food cost",
    ),
    "Hummingbirds": ("Common name", "Scientific name", "Group", "Benefit", "Beak direction"),
    "Bonus": ("Bonus card", "Set", "Condition"),
    "Goals": ("Goal", "Set", "Reverse"),
}

EXPECTED_COLUMNS = {
    "Birds": BIRD_COLUMNS,
    "Hummingbirds": HUMMINGBIRD_COLUMNS,
    "Bonus": BONUS_COLUMNS,
    "Goals": GOAL_COLUMNS,
}

SET_TO_CONTENT_PACK = {
    "core": ContentPack.CORE,
    "european": ContentPack.EUROPEAN,
    "oceania": ContentPack.OCEANIA,
    "asia": ContentPack.ASIA,
    "americas": ContentPack.AMERICAS,
    "promoUS": ContentPack.PROMO_US,
    "promoEurope": ContentPack.PROMO_EUROPE,
    "promoAsia": ContentPack.PROMO_ASIA,
    "promoNZ": ContentPack.PROMO_NZ,
    "promoCA": ContentPack.PROMO_CA,
    "promoUK": ContentPack.PROMO_UK,
}


@dataclass(frozen=True)
class FieldIssue:
    """A field-level workbook issue found during audit."""

    sheet: str
    row_number: int | None
    field: str
    message: str
    value: Any = None


@dataclass
class SheetAudit:
    """Audit summary for one workbook sheet."""

    name: str
    row_count: int
    column_count: int
    missing_columns: list[Any] = field(default_factory=list)
    extra_columns: list[Any] = field(default_factory=list)
    non_empty_counts: dict[Any, int] = field(default_factory=dict)
    value_counts: dict[str, Counter[str]] = field(default_factory=dict)


@dataclass
class WorkbookAudit:
    """Complete audit summary for a workbook."""

    path: Path
    sheets: dict[str, SheetAudit]
    ignored_sheets: list[str]
    issues: list[FieldIssue]
    expansion_coverage: dict[str, Counter[str]]


def is_blank(value: Any) -> bool:
    """Return true for empty spreadsheet values."""

    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_cell(value: Any) -> Any:
    """Normalize spreadsheet blanks without changing meaningful values."""

    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def read_sheet_rows(path: Path, sheet_name: str) -> tuple[list[Any], list[dict[Any, Any]]]:
    """Read non-empty rows from one workbook sheet."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = [normalize_cell(value) for value in next(rows_iter)]
    rows: list[dict[Any, Any]] = []

    key_header = headers[0]
    for row_number, values in enumerate(rows_iter, start=2):
        normalized_values = [normalize_cell(value) for value in values]
        if all(is_blank(value) for value in normalized_values):
            continue
        row = dict(zip(headers, normalized_values, strict=False))
        if is_blank(row.get(key_header)):
            continue
        row["__row_number__"] = row_number
        rows.append(row)

    return headers, rows


def audit_workbook(path: str | Path) -> WorkbookAudit:
    """Audit workbook sheets, expected columns, field gaps, and value domains."""

    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    ignored_sheets = [name for name in workbook.sheetnames if name not in EXPECTED_SHEETS]

    sheet_audits: dict[str, SheetAudit] = {}
    issues: list[FieldIssue] = []
    expansion_coverage: dict[str, Counter[str]] = {}

    for sheet_name in EXPECTED_SHEETS:
        headers, rows = read_sheet_rows(workbook_path, sheet_name)
        expected_columns = list(EXPECTED_COLUMNS[sheet_name])
        sheet_audit = SheetAudit(
            name=sheet_name,
            row_count=len(rows),
            column_count=len(headers),
            missing_columns=[column for column in expected_columns if column not in headers],
            extra_columns=[column for column in headers if column not in expected_columns],
        )

        for column in headers:
            if column is None:
                continue
            non_empty = sum(1 for row in rows if not is_blank(row.get(column)))
            sheet_audit.non_empty_counts[column] = non_empty

        for row in rows:
            row_number = int(row["__row_number__"])
            for column in REQUIRED_COLUMNS[sheet_name]:
                if is_blank(row.get(column)):
                    issues.append(
                        FieldIssue(
                            sheet=sheet_name,
                            row_number=row_number,
                            field=str(column),
                            message="required source field is blank",
                        )
                    )

        sheet_audits[sheet_name] = sheet_audit

    _audit_bird_domains(workbook_path, sheet_audits["Birds"], issues)
    _audit_bonus_domains(workbook_path, issues)
    _audit_goal_domains(workbook_path, issues)

    for sheet_name in ("Birds", "Bonus", "Goals"):
        _, rows = read_sheet_rows(workbook_path, sheet_name)
        expansion_coverage[sheet_name] = _count_content_packs(rows)

    return WorkbookAudit(
        path=workbook_path,
        sheets=sheet_audits,
        ignored_sheets=ignored_sheets,
        issues=issues,
        expansion_coverage=expansion_coverage,
    )


def _audit_bird_domains(path: Path, sheet_audit: SheetAudit, issues: list[FieldIssue]) -> None:
    _, rows = read_sheet_rows(path, "Birds")
    domain_columns = {
        "Set": set(SET_TO_CONTENT_PACK),
        "Color": {"brown", "white", "pink", "teal", "yellow"},
        "Nest type": {"ground", "bowl", "cavity", "platform", "wild"},
        "Beak direction": {"L", "R", "N"},
    }
    value_count_columns = ("Set", "Color", "Nest type", "Beak direction")

    for column in value_count_columns:
        sheet_audit.value_counts[column] = Counter(
            str(row[column]) for row in rows if not is_blank(row.get(column))
        )

    for row in rows:
        row_number = int(row["__row_number__"])
        for column, allowed_values in domain_columns.items():
            value = row.get(column)
            if not is_blank(value) and value not in allowed_values:
                issues.append(
                    FieldIssue(
                        sheet="Birds",
                        row_number=row_number,
                        field=column,
                        value=value,
                        message=(
                            "unexpected domain value; add normalization mapping "
                            "or review source"
                        ),
                    )
                )

        for column, message in {
            "Color": (
                "blank power color; normalize to none only after confirming "
                "the card has no power"
            ),
            "Nest type": "blank nest type; normalize to explicit special/non-nesting category",
        }.items():
            if is_blank(row.get(column)):
                issues.append(
                    FieldIssue(
                        sheet="Birds",
                        row_number=row_number,
                        field=column,
                        value=row.get("Common name"),
                        message=message,
                    )
                )

        wingspan = row.get("Wingspan")
        if not is_blank(wingspan) and not isinstance(wingspan, int | float):
            issues.append(
                FieldIssue(
                    sheet="Birds",
                    row_number=row_number,
                    field="Wingspan",
                    value=wingspan,
                    message="non-numeric wingspan; normalize as variable or unknown wingspan",
                )
            )


def _audit_bonus_domains(path: Path, issues: list[FieldIssue]) -> None:
    _, rows = read_sheet_rows(path, "Bonus")
    for row in rows:
        row_number = int(row["__row_number__"])
        if is_blank(row.get("VP")):
            issues.append(
                FieldIssue(
                    sheet="Bonus",
                    row_number=row_number,
                    field="VP",
                    value=row.get("Bonus card"),
                    message="bonus scoring text is blank; likely needs hand-authored scoring rule",
                )
            )


def _audit_goal_domains(path: Path, issues: list[FieldIssue]) -> None:
    _, rows = read_sheet_rows(path, "Goals")
    for row in rows:
        row_number = int(row["__row_number__"])
        missing_scoring = [column for column in (1, 2, 3, 4) if is_blank(row.get(column))]
        if missing_scoring:
            issues.append(
                FieldIssue(
                    sheet="Goals",
                    row_number=row_number,
                    field="round_end_scoring",
                    value=row.get("Goal"),
                    message="placement scoring columns are blank; likely duet/map goal",
                )
            )


def _count_content_packs(rows: list[dict[Any, Any]]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for row in rows:
        set_value = row.get("Set")
        if is_blank(set_value):
            continue
        for raw_set in str(set_value).split(","):
            cleaned = raw_set.strip()
            content_pack = SET_TO_CONTENT_PACK.get(cleaned)
            counts[content_pack.value if content_pack else cleaned] += 1
    return counts


def format_markdown_report(audit: WorkbookAudit) -> str:
    """Format an audit summary as Markdown."""

    lines = [
        "# Wingspan Card Workbook Audit",
        "",
        f"Source workbook: `{audit.path.name}`",
        "",
        "## Sheets",
        "",
        "| Sheet | Rows | Columns | Missing expected columns | Extra source columns |",
        "|---|---:|---:|---|---|",
    ]

    for sheet in audit.sheets.values():
        missing = ", ".join(str(value) for value in sheet.missing_columns) or "None"
        extra = ", ".join(str(value) for value in sheet.extra_columns) or "None"
        lines.append(
            f"| {sheet.name} | {sheet.row_count} | {sheet.column_count} | "
            f"{missing} | {extra} |"
        )

    lines.extend(["", "Ignored workbook artifact sheets: " + ", ".join(audit.ignored_sheets), ""])
    lines.extend(["## Expansion Coverage", ""])

    for sheet_name, counts in audit.expansion_coverage.items():
        lines.append(f"### {sheet_name}")
        lines.append("")
        lines.append("| Content pack | Rows |")
        lines.append("|---|---:|")
        for content_pack, count in sorted(counts.items()):
            lines.append(f"| {content_pack} | {count} |")
        lines.append("")

    lines.extend(["## Field Issues", ""])
    if not audit.issues:
        lines.append("No field issues found.")
    else:
        issue_counts = Counter(
            f"{issue.sheet}:{issue.field}:{issue.message}" for issue in audit.issues
        )
        lines.append("| Count | Sheet | Field | Issue |")
        lines.append("|---:|---|---|---|")
        for key, count in issue_counts.most_common():
            sheet, field_name, message = key.split(":", maxsplit=2)
            lines.append(f"| {count} | {sheet} | {field_name} | {message} |")

    lines.extend(["", "## Normalization Needs", ""])
    lines.extend(_format_normalization_notes(audit))
    lines.append("")
    return "\n".join(lines)


def _format_normalization_notes(audit: WorkbookAudit) -> list[str]:
    bird_sheet = audit.sheets["Birds"]
    notes = [
        "- Map workbook set labels such as `promoUS` and `promoEurope` "
        "to snake-case content packs.",
        "- Treat `Color` blanks as `PowerColor.NONE` only after confirming "
        "those birds have no power.",
        "- Treat blank `Nest type` values as an explicit non-nesting or special "
        "nest category decision.",
        "- Convert `Wingspan` value `*` to `wingspan_is_variable=true` with `wingspan_cm=null`.",
        "- Convert `Beak direction` values `L`, `R`, and `N` to normalized enum values.",
        "- Decide whether multi-direction values such as `LR` and `LL` are valid "
        "Asia metadata or source errors.",
        "- Parse `VP` scoring text on bonus cards into hand-authored scoring handlers over time.",
        "- Split duet/map goals from standard end-of-round goals because they do "
        "not use columns `1` through `4`.",
    ]

    set_counts = bird_sheet.value_counts.get("Set")
    if set_counts:
        notes.append(
            "- Bird rows cover these workbook sets: "
            + ", ".join(f"`{name}` ({count})" for name, count in sorted(set_counts.items()))
            + "."
        )
    return notes


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit the Wingspan card workbook.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default="data/raw/wingspan-card-list.xlsx",
        help="Path to the workbook to audit.",
    )
    args = parser.parse_args()
    audit = audit_workbook(args.workbook)
    print(format_markdown_report(audit))


if __name__ == "__main__":
    main()
